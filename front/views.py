import json
import logging
import os
import traceback

from django.contrib.messages.storage import session
from django.http import HttpResponse
from django.shortcuts import render, render_to_response, redirect

# Create your views here.
from django.template import RequestContext
from django.urls import reverse

import openpyxl

logger = logging.getLogger(__name__)


def index(request):
    '''
    上传文件首页
    :param request:
    :return:html
    '''
    results = dict()
    results['name'] = '66666'
    return render(request, 'upload.html', results)


# 文件上传函数
def file(request):
    '''
    文件上传
    :param request:
    :return:
    '''
    results = dict()
    # assert False
    if request.method == "POST":
        try:
            myFile = request.FILES.get("myfile", None)  # 获取上传的文件，如果没有文件，则默认为None
            # file_name = myFile.split('.')[-2]
            if not myFile:
                results['error_msg'] = '你没有上传文件啊!'
                raise Exception('你没有上传文件啊')
            baseDir = os.path.dirname(os.path.abspath(__name__));
            jpgdir = os.path.join(baseDir, 'static', 'files');
            dir_path = os.path.join(jpgdir, myFile.name)
            destination = open(dir_path, 'wb+')  # 打开特定的文件进行二进制的写操作
            for chunk in myFile.chunks():  # 分块写入文件
                destination.write(chunk)
            destination.close()
            print(dir_path)
            print(type(dir_path))
            return redirect(reverse('front:test', args=[myFile.name]), )
        except Exception as e:
            print(e)
            if 'error_msg' not in results:
                results['error_msg'] = e
            return render(request, 'error.html', results)


def test(request, path):
    results = dict()
    print(path)
    results['path'] = path
    return render(request, 'test.html', results)


def go(request):
    '''
    调用梅雨写的函数
    :param request:
    :return:
    '''
    results = dict()
    try:
        results['is_succ'] = True
    except Exception as e:
        results['is_succ'] = False
    return HttpResponse(json.dumps(results), content_type="application/json")


def error(request):
    '''

    :param request:
    :return:
    '''
    return render(request, 'error.html')


# 从这里开始
import pandas as pd
import numpy as np
import datetime as dt
import os
import openpyxl as xl

pd.set_option("Display.max_columns", None)
pd.set_option("Display.max_rows", None)
SemanticTag = "SemanticTag1"
DocumentType = "DocumentType1"
EncounterType = "EncounterType1"
DocumentName = "DocumentName1"
GlobalEMPI = "GlobalEMPI"
GlobalTimeStamp = "GlobalTimeStamp1"


def blank_record(toSheet):
    if toSheet == "病案首页":
        Record = pd.DataFrame(
            columns=["EMPI", "主要诊断", "主诊断转归", "住院号", "住院天数", "入院日期", "出生日期", "出院日期", "出院科室", "卡号", "地址",
                     "就诊医院", "影像号",
                     "治疗号", "电话", "病人姓名", "病人年龄", "病人性别", "联系人地址", "联系人姓名", "联系人电话", "身份证"])
    elif toSheet == "病史记录":
        Record = pd.DataFrame(columns=["EMPI", "住院号", "病人类型", "报告时间", "模板名称", "文书内容"])
    elif toSheet == "影像学报告":
        Record = pd.DataFrame(columns=["EmpiId", "住院号", "病人类型", "报告日期", "报告类型", "检查部位", "文书"])
    elif toSheet == "检验结果":
        Record = pd.DataFrame(
            columns=["EMPI", "住院号", "病人类型", "报告时间", "检验项目编码", "检验项目名称", "数值结果", "文本结果", "备注", "备注1",
                     "单位", "参考值"])
    return Record


def get_timestamp_col(df):
    timestamp = list(set(df["GlobalTimeStamp1"].dropna()))[0]
    return timestamp


def get_empi_col(df):
    empi = list(set(df["GlobalEMPI"].dropna()))[0]
    return empi


def get_data(path):
    file = pd.read_excel(path)
    # 这是file
    start_row = file.loc[file["Heads"] == "data"].index.to_list()[0]  # ？？需要问的
    data = file.iloc[start_row + 1:]
    data.replace(" ", None)
    data.replace()
    return data


def front_pages(table, path):
    """
    查找病案首页数据，生成病案首页，只适用于一行一次入院的情况
    :param table:
    :return: frontpage
    """
    print("front page start")
    front_page_out = blank_record("病案首页")
    front_page = table.loc[(table["DocumentType1"] == "病案首页") & (table["SemanticTag1"] != "TimeStamp")]
    if not front_page.empty:
        date_column_name = get_timestamp_col(front_page)
        EMPI_name = get_empi_col(front_page)
        data = get_data(path)
        columns = front_page.index
        for i in data.index:
            record = blank_record("病案首页")
            record.loc[i, "入院日期"] = data.loc[i, date_column_name]
            record.loc[i, "EMPI"] = data.loc[i, EMPI_name]
            for c in columns:
                according_column = front_page.loc[c, "DocumentName1"]
                if according_column in list(record.columns):
                    record.loc[i, according_column] = data.loc[i, c]
                else:
                    print("数据与模板不符", according_column)
            # record.loc[i,"住院天数"] =
            front_page_out = front_page_out.append(record)
    print("front page end")
    return front_page_out


def get_out_patient_date(df, in_date):
    try:
        out_patient = df.loc[(df[DocumentName] == "出院日期") & (df[GlobalTimeStamp] == in_date)].index.to_list()[0]
    except Exception as e:
        out_patient = None
    return out_patient


def multiple_front_page(table, path):
    # print("front page start")
    front_page_out = blank_record("病案首页")
    date_column_names = list(set(table["GlobalTimeStamp1"].dropna()))
    data = get_data(path)
    front_page = table.loc[(table["DocumentType1"] == "病案首页") & (table["SemanticTag1"] != "TimeStamp")]
    for i in data.index:
        for date_column_name in date_column_names:
            # print(date_column_name)
            if not front_page.empty:
                columns = front_page.index
                EMPI_name = get_empi_col(front_page)
                record = blank_record("病案首页")
                date = data.loc[i, date_column_name]
                if not pd.isnull(date) and not str(date).isspace():
                    record.loc[i, "入院日期"] = data.loc[i, date_column_name]
                    record.loc[i, "EMPI"] = data.loc[i, EMPI_name]
                    out_patient = get_out_patient_date(table, date_column_name)
                    if out_patient:
                        record.loc[i, "出院日期"] = data.loc[i, get_out_patient_date(table, date_column_name)]
                    for c in columns:
                        according_column = front_page.loc[c, "DocumentName1"]
                        if according_column in list(record.columns):
                            record.loc[i, according_column] = data.loc[i, c]
                        else:
                            print("数据与模板不符", according_column)
                    # record.loc[i,"住院天数"] =
                    # print(record)
                    # print(front_page_out)
                    front_page_out = front_page_out.append(record)
    print("front page end")
    return front_page_out


def three_pages(table, path):
    print("other pages start")
    medical_records = blank_record("病史记录")
    radio_records = blank_record("影像学报告")
    test_records = blank_record("检验结果")
    other_page = table.loc[table["DocumentType1"] != "病案首页"]
    # 先按照timestamp分组
    x = other_page.groupby(other_page["GlobalTimeStamp1"]).groups.keys()
    # 逐个timestamp处理
    for one_date in x:
        cols = other_page.loc[table["GlobalTimeStamp1"] == one_date]
        # 找记录类型列 病案首页 病史记录 影像学报告 检验结果
        sheet_name = cols.groupby(cols["DocumentName1"]).groups.keys()
        # 找出同一timestamp下的同记录类型，即为输出表中的一行
        # 逐个DocumentName1处理
        for one_record in sheet_name:
            # 对应的时间戳列
            date_col = table.loc[one_date]
            # 同记录类型
            record_cols = cols.loc[cols["DocumentName1"] == one_record]
            empi_col = get_empi_col(record_cols)
            record_type = record_cols["DocumentName1"]
            # 对于有structure的数据，修改时间戳列的记录类型
            stucture_flag = 0
            if "structured" in record_type[0]:
                stucture_flag = 1
                record_name = record_type[0].split("/")[0]
                date_col["DocumentName1"] = record_name
            # 找记录所属的sheet，如果出现一个以上结果，说明有标注错误
            sheet_type = list(set(record_cols["DocumentType1"]))
            if len(sheet_type) != 1:
                print("DocumentType冲突")
            toSheet = sheet_type[0]
            # 开始按照规则处理数据
            data = get_data(path)
            for i in data.index:
                print(i, len(data))
                # 如果时间为空，跳过改行
                date = data.loc[i, date_col.name]
                if not pd.isnull(date) and not str(date).isspace():
                    record = blank_record(toSheet)
                    finding = ""
                    record.loc[i, record.columns[3]] = data.loc[i, date_col.name]
                    record.loc[i, record.columns[2]] = date_col["EncounterType1"]
                    empi = data.loc[i, empi_col]
                    record.loc[i, record.columns[0]] = empi
                    # record.loc[i, record.columns[2]] = list(front_page.loc[front_page["EMPI"] == empi, "住院号"])[0]
                    for col_name in record_cols.index:
                        if stucture_flag == 1:
                            finding = finding + " %s：%s" % (col_name, data.loc[i, col_name])
                        elif len(record_cols) != 1:
                            finding = finding + "%s：%s" % (
                            record_cols.loc[col_name, "DocumentName1"], data.loc[i, col_name])
                        else:
                            finding = data.loc[i, col_name]
                    if toSheet == "病史记录":
                        record.loc[i, "模板名称"] = date_col["DocumentName1"]
                        record.loc[i, "文书内容"] = finding
                        medical_records = medical_records.append(record)
                        # print (medical_records)
                    elif toSheet == "影像学报告":
                        record.loc[i, "报告类型"] = date_col["DocumentName1"]
                        record.loc[i, "文书"] = finding
                        radio_records = radio_records.append(record)
                    elif toSheet == "检验结果":
                        record.loc[i, "检验项目名称"] = date_col["DocumentName1"]
                        value_type = list(record_cols["DocumentName1"])[0]
                        record.loc[i, value_type] = list(data.loc[i, record_cols.index])[0]
                        test_records = test_records.append(record)
    print("other pages end")
    return medical_records, radio_records, test_records


def generate(path, type="multiple"):
    file = pd.read_excel(path, index_col="Heads")
    file.replace(" ", None)
    try:
        table = file.iloc[0:6].drop("end", axis=1).transpose()  # .reset_index()
    except Exception as e:
        table = file.iloc[0:6].transpose()
    if type == "oneline":
        front_page = front_pages(table, path)
    elif type == "multiple":
        front_page = multiple_front_page(table, path)
    medical_records, radio_records, test_records = three_pages(table, path)
    return front_page, medical_records, radio_records, test_records


def quality_control(front_page, medical_records, radio_records, test_records, default=None):
    for i in front_page.index:
        print(i)
        if pd.isnull(front_page.loc[i, "入院日期"]):
            front_page = front_page.drop(i)
            continue
        elif pd.isnull(front_page.loc[i, "出院日期"]):
            front_page.loc[i, "出院日期"] = front_page.loc[i, "入院日期"] + dt.timedelta(days=14)
        elif (front_page.loc[i, "出院日期"] - front_page.loc[i, "入院日期"]).days < 0:
            print(i, "con")
            front_page = front_page.drop(i)
            continue
        if pd.isnull(front_page.loc[i, "出生日期"]):
            print(i)
            try:
                front_page.loc[i, "出生日期"] = front_page.loc[front_page["EMPI"] == front_page.loc[
                    i, "EMPI"], "入院日期"].min() - dt.timedelta(days=int(front_page.loc[i, "病人年龄"]) * 365)
            except Exception as e:
                print(e)
        if default:
            column = default[0]
            value = default[1]
            front_page.loc[i, column] = value
    return front_page, medical_records, radio_records, test_records


def format_front(front_page):
    for i in front_page.index:
        # front_page.loc[i,"入院日期"] = dt.datetime.strptime(front_page.loc[i,"入院日期"],"%Y-%m-%d")
        # front_page.loc[i,"出院日期"] = dt.datetime.strptime(front_page.loc[i,"出院日期"],"%Y-%m-%d %H:%M:%S")

        if not pd.isnull(front_page.loc[i, "入院日期"]) and not pd.isnull(front_page.loc[i, "出院日期"]):
            front_page.loc[i, "住院天数"] = (front_page.loc[i, "出院日期"] - front_page.loc[i, "入院日期"]).days
        if not pd.isnull(front_page.loc[i, "入院日期"]):
            front_page.loc[i, "入院日期"] = dt.datetime.strftime(front_page.loc[i, "入院日期"], "%Y%m%d%H%M%S-000")
        if not pd.isnull(front_page.loc[i, "出院日期"]):
            front_page.loc[i, "出院日期"] = dt.datetime.strftime(front_page.loc[i, "出院日期"], "%Y%m%d%H%M%S-000")
        if not pd.isnull(front_page.loc[i, "出生日期"]):
            front_page.loc[i, "出生日期"] = dt.datetime.strftime(front_page.loc[i, "出生日期"], "%Y-%m-%d")
    return front_page


def format_page(some_records, front_page):
    for i in some_records.index:
        date = some_records.loc[i, some_records.columns[3]]
        if isinstance(date, str):
            if "/" in date:
                some_records.loc[i, some_records.columns[3]] = dt.datetime.strptime(
                    some_records.loc[i, some_records.columns[3]], "%Y/%m/%d %H:%M:%S")
            elif ":" in date:
                some_records.loc[i, some_records.columns[3]] = dt.datetime.strptime(
                    some_records.loc[i, some_records.columns[3]], "%Y-%m-%d %H:%M:%S")
            else:
                some_records.loc[i, some_records.columns[3]] = dt.datetime.strptime(
                    some_records.loc[i, some_records.columns[3]], "%Y-%m-%d")
        some_records.loc[i, some_records.columns[3]] = dt.datetime.strftime(
            some_records.loc[i, some_records.columns[3]], "%Y%m%d%H%M%S-000")
        # else:
        #     some_records.loc[i, some_records.columns[4]] = dt.datetime.strftime(dt.datetime.strptime(
        #         some_records.loc[i, some_records.columns[4]], "%Y-%m-%d"), "%Y%m%d%H%M%S-000")
        try:
            empi = some_records.loc[i, some_records.columns[0]]
            zyh = front_page.loc[front_page["EMPI"] == empi, "住院号"]
            some_records.loc[i, some_records.columns[1]] = str(int(list(zyh)[0]))
        except:
            continue
        # print(i,len(some_records))
    return some_records


def format_table(front_page, medical_records, radio_records, test_records):
    try:
        front_page = format_front(front_page.reset_index(drop=True))
        medical_records = format_page(medical_records.reset_index(drop=True), front_page.reset_index(drop=True))
        radio_records = format_page(radio_records.reset_index(drop=True), front_page.reset_index(drop=True))
        test_records = format_page(test_records.reset_index(drop=True), front_page.reset_index(drop=True))
    except Exception as e:
        print(e)
    return front_page, medical_records, radio_records, test_records


def group_by_empi(front_page, medical_records, radio_records, test_records, EMPI):
    front = front_page.loc[front_page[front_page.columns[0]] == EMPI]
    medical = medical_records.loc[medical_records[medical_records.columns[0]] == EMPI]
    radio = radio_records.loc[radio_records[radio_records.columns[0]] == EMPI]
    test = test_records.loc[test_records[test_records.columns[0]] == EMPI]
    return front, medical, radio, test


def output(front_page, medical_records, radio_records, test_records, outputdir):
    EMPIs = list(set(front_page["EMPI"]))
    data = list()
    for EMPI in EMPIs:
        front, medical, radio, test = group_by_empi(front_page, medical_records, radio_records, test_records, EMPI)
        # front, medical, radio, test = quality_control(front, medical, radio, test)
        outpath = os.path.join(outputdir, str(EMPI) + ".xlsx")
        datapath = os.path.join('static/excels',str(EMPI)+".xlsx")
        print(EMPI, outpath)
        with pd.ExcelWriter(outpath) as writer:
            front.to_excel(writer, "病案首页", index=False)
            medical.to_excel(writer, "病史记录", index=False)
            radio.to_excel(writer, "影像学报告", index=False)
            test.to_excel(writer, "检验结果", index=False)
        data.append(datapath)
    return data


def meiyu(request):
    print('*' * 30)
    # path = "data/rawdata/Demo.xlsx"
    results = dict()
    try:
        path1 = request.POST.get('path')
        # outputdir = "data/batch/demo_output"
        baseDir = os.path.dirname(os.path.abspath(__name__));
        jpgdir = os.path.join(baseDir, 'static', 'files');
        outputdir = os.path.join(baseDir, 'static', 'excels');
        # outputdir = "/static/excels"
        path = os.path.join(jpgdir, path1)
        front_page, medical_records, radio_records, test_records = generate(path, "multiple")
        front_page, medical_records, radio_records, test_records = quality_control(front_page.reset_index(drop=True),
                                                                                   medical_records.reset_index(
                                                                                       drop=True),
                                                                                   radio_records.reset_index(drop=True),
                                                                                   test_records.reset_index(drop=True),
                                                                                   default=["主要诊断", "直肠癌"])
        front_page, medical_records, radio_records, test_records = format_table(front_page, medical_records,
                                                                                radio_records, test_records)
        data = output(front_page, medical_records, radio_records, test_records, outputdir)
        # print(data)
        results['is_succ'] = True
        results['data'] = data
    except Exception as e:
        results['is_succ'] = False
    return HttpResponse(json.dumps(results), content_type="application/json")
    #     return render(request,'deal_error.html')
    # return render(request,'deal_success.html',data)


def go_on(request):
    return render(request, 'digist.html')


def sheetname(request):
    wb = openpyxl.load_workbook('static/excels/1.xlsx')
    print(wb.sheetnames)
    ws = wb.active
    print(ws.title)
    print('{} * {}'.format(ws.max_row, ws.max_column))
    # sheet3 = wb.get_sheet_by_name('检验结果')
    # sheet3 = wb.['检验结果']
    # getting cell  from the sheets
    # print(ws['A1'])
    # print(ws['A1'].value)
    # c = ws['B1']
    # print('Row {} Col {} is {}'.format(c.row,c.column,c.value))
    # print('Cell {} is {}{'.format(c.coordinate,c.value))
    results = dict()
    # for i in range(ws.max_row):
    #     for j in range(ws.max_column):
    #         # if ws.cell(row=i+1,column=j+1).value != None:
    #         print(ws.cell(row=i+1,column=j+1).value)
    #         results['{}_{}'.format(i+1,j+1)] = ws.cell(row=i+1,column=j+1).value
    # print('\n')
    data_list = list()
    data = dict()
    for i in range(ws.max_row):
        dic = 'data_{}'.format(i)
        dic = dict()
        for j in range(ws.max_column):
            # data[j] = ws.cell(row=i+1,column=j+1).value
            # data_name = '{}_{}'.format(i,j)
            # data_name = dict()
            # data_name[j] = ws.cell(row=i+1,column=j+1).value
            # results['j'] = data_name
            # results['{}_{}'.format(i,j)] = ws.cell(row=i+1,column=j+1).value
            dic[j] = ws.cell(row=i + 1, column=j + 1).value

        data_list.append(dic)
    return render(request, 'table.html', {'data_list': data_list})



def del_file(request):
    '''
    删除文件命令
    :param request:
    :return:
    '''
    basedir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    abs_path = os.path.join(basedir, 'static\excels')
    fileslist = []
    fileslist = os.listdir(abs_path)
    print(fileslist)
    for file in fileslist:
        filepath = os.path.join(abs_path,file)
        os.remove(filepath)
    return  HttpResponse('ok')



def text(request):
    return  render(request,'index/index.html')

from datetime import  *
import time
def tim(request):
    '''
    时间函数
    :param request:
    :return:
    '''
    start_time = datetime.now()
    time.sleep(10)
    end_time = datetime.now()
    del_time = (end_time - start_time).total_seconds()
    for i in range(10):
        print(i)
    return  HttpResponse(del_time)
